"""
SCF Checker v6.0
Nokia BTS Configuration Validator
- Vertical button layout
- 3-panel results bottom
- RTPOL reads srcIP directly from SCF IPIF objects
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import xml.etree.ElementTree as ET
import os, sys, re, json

try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# ══════════════════════════════════════════
#  COLORS
# ══════════════════════════════════════════
BG_WIN    = "#f0f2f5"
BG_PANEL  = "#ffffff"
BG_CARD   = "#f7f8fa"
BG_INPUT  = "#ffffff"
BORDER    = "#d0d7e2"
ACCENT    = "#0066cc"
ACCENT_LT = "#e8f0fb"
GREEN     = "#1a7a3c"
GREEN_BG  = "#e8f5ec"
RED       = "#cc2200"
RED_BG    = "#fdf0ee"
YELLOW    = "#7a6000"
YELLOW_BG = "#fffbe6"
ORANGE    = "#b35a00"
ORANGE_BG = "#fff4e5"
TEXT1     = "#1a2233"
TEXT2     = "#5a6a7a"
TEXT3     = "#9aaabb"
WHITE     = "#ffffff"

F_HEAD  = ("Segoe UI", 10, "bold")
F_BODY  = ("Segoe UI", 10)
F_SMALL = ("Segoe UI", 9)
F_MONO  = ("Consolas", 9)
F_NUM   = ("Segoe UI", 8, "bold")

REGIONS   = ["CR", "NR", "ER"]
SOLUTIONS = [
    "2G","3G","2G-3G (SRAN)","2G-3G-FDD (SRAN)",
    "2G-FDD (SRAN)","2G-FDD-5G (SRAN)","3G-FDD (SRAN)",
    "FDD","TDD","5G","TDD-5G (SRAN)",
]

# ══════════════════════════════════════════
#  REQUIREMENTS PER SOLUTION/REGION
# ══════════════════════════════════════════
SOLUTION_REQUIREMENTS = {
    ("ER", "TDD-5G (SRAN)"): [
        {"id":"check_scf",    "label":"Check the SCF file",
         "auto":True,  "trigger":"run"},
        {"id":"add_iprt1",    "label":"Add the missing relations in IPRT-1",
         "auto":False, "trigger":"write", "checked":True},
        {"id":"delete_iprt7", "label":"Delete the relation duplication of TDD (IPRT-7)",
         "auto":False, "trigger":"write", "checked":False},
        {"id":"add_rtpol",    "label":"Add missing Routing Policies (IPRT-8 LTE-Traffic & IPRT-9 LTE-Signalling)",
         "auto":False, "trigger":"write", "checked":False},
    ],
}

def _get_save_path():
    try:
        base = os.path.dirname(os.path.abspath(__file__))
        path = os.path.join(base, "scf_requirements.json")
        open(path, "a").close()
        return path
    except Exception:
        return os.path.join(os.path.expanduser("~"), "scf_requirements.json")

def save_requirements():
    try:
        data = {}
        for (region, sol), reqs in SOLUTION_REQUIREMENTS.items():
            data[f"{region}||{sol}"] = [{k: v for k, v in r.items()} for r in reqs]
        with open(_get_save_path(), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True, "Saved."
    except Exception as e:
        return False, str(e)

def load_requirements():
    path = _get_save_path()
    if not os.path.exists(path):
        return
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        SOLUTION_REQUIREMENTS.clear()
        for key_str, reqs in data.items():
            if "||" not in key_str: continue
            region, sol = key_str.split("||", 1)
            SOLUTION_REQUIREMENTS[(region, sol)] = reqs
    except Exception:
        pass

def get_requirements(region, solution):
    key = (region, solution)
    if key in SOLUTION_REQUIREMENTS:
        return SOLUTION_REQUIREMENTS[key]
    for k, v in SOLUTION_REQUIREMENTS.items():
        if k[1] == solution:
            return v
    return []

load_requirements()


# ══════════════════════════════════════════
#  EXCEL READER
# ══════════════════════════════════════════
def read_relations_from_excel(excel_path, sheet_name):
    if not EXCEL_SUPPORT:
        messagebox.showerror("Error", "openpyxl not installed.")
        return []
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        messagebox.showerror("Excel Error", str(e))
        return []

    matched = None
    for sn in wb.sheetnames:
        if sn.strip().upper() == sheet_name.strip().upper():
            matched = sn
            break
    if not matched:
        messagebox.showwarning("Sheet Not Found",
            f"Sheet '{sheet_name}' not found.\nAvailable: {', '.join(wb.sheetnames)}")
        wb.close()
        return []

    ws = wb[matched]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows: return []

    header = all_rows[0]
    header_str = [str(h).strip().lower().replace(" ","").replace("_","") if h else "" for h in header]

    has_dest = any(h in header_str for h in ["dest","destipaddr","destination"])
    has_pfx  = any(h in header_str for h in ["prefix","destinationipprefixlength","prefixlength","mask"])

    if has_dest and has_pfx:
        def col(variants):
            for v in variants:
                if v in header_str: return header_str.index(v)
            return None
        idx_iprt   = col(["iprt","iprtid","routetable"])
        idx_dest   = col(["dest","destipaddr","destination"])
        idx_prefix = col(["prefix","destinationipprefixlength","prefixlength","mask"])
        idx_gw     = col(["gateway","gw","nexthop"])
        idx_pref   = col(["preference","pref","metric"])
        relations = []
        for row in all_rows[1:]:
            def get(idx):
                if idx is None: return ""
                v = row[idx] if idx < len(row) else None
                return str(v).strip() if v is not None and str(v).strip() not in ("None","") else ""
            dest = get(idx_dest)
            if not dest: continue
            relations.append({"iprt":get(idx_iprt),"dest":dest,"prefix":get(idx_prefix),
                               "gateway":"FROM_SCF","pref":get(idx_pref) or "1"})
        return relations
    else:
        # Nokia SBR format — cols W/X
        dest_col, mask_col = None, None
        for i, h in enumerate(header):
            if h:
                hs = str(h).strip()
                if "Static Route" in hs or "IPRT-1 Static" in hs: dest_col = i
                elif hs == "Mask": mask_col = i
        if dest_col is None: dest_col, mask_col = 22, 23
        relations = []
        for row in all_rows[1:]:
            dest = row[dest_col] if dest_col < len(row) else None
            mask = row[mask_col] if mask_col and mask_col < len(row) else None
            if not dest or str(dest).strip() in ("","None"): continue
            dest_str = str(dest).strip()
            if not re.match(r'^\d+\.\d+\.\d+\.\d+$', dest_str): continue
            mask_str = str(int(mask)) if mask and str(mask).strip() not in ("","None") else "0"
            relations.append({"iprt":"IPRT-1","dest":dest_str,"prefix":mask_str,
                               "gateway":"FROM_SCF","pref":"1"})
        return relations


# ══════════════════════════════════════════
#  XML HELPERS
# ══════════════════════════════════════════
def _tag(el):
    return el.tag.split("}")[-1] if "}" in el.tag else el.tag

def parse_scf_iprt1(xml_path):
    tree = ET.parse(xml_path)
    root = tree.getroot()
    for mo in root.iter():
        if _tag(mo) != "managedObject": continue
        if not mo.get("distName","").endswith("IPRT-1"): continue
        routes = []
        for lst in mo.iter():
            if _tag(lst) != "list" or lst.get("name") != "staticRoutes": continue
            for item in lst:
                if _tag(item) != "item": continue
                e = {p.get("name"): (p.text or "").strip()
                     for p in item if _tag(p) == "p"}
                if e:
                    routes.append({"dest":e.get("destIpAddr",""),
                                   "prefix":e.get("destinationIpPrefixLength",""),
                                   "gateway":e.get("gateway",""),
                                   "pref":e.get("preference","1")})
        return routes
    return []

def get_gateway_from_scf(xml_path):
    tree = ET.parse(xml_path)
    root = tree.getroot()
    for mo in root.iter():
        if _tag(mo) != "managedObject": continue
        if not mo.get("distName","").endswith("IPRT-1"): continue
        for lst in mo.iter():
            if _tag(lst) != "list" or lst.get("name") != "staticRoutes": continue
            for item in lst:
                for p in item:
                    if _tag(p) == "p" and p.get("name") == "gateway":
                        gw = (p.text or "").strip()
                        if gw and gw != "0.0.0.0": return gw
    return "0.0.0.0"

def get_distbase(xml_path):
    tree = ET.parse(xml_path)
    root = tree.getroot()
    for mo in root.iter():
        if _tag(mo) == "managedObject":
            dn = mo.get("distName","")
            if "MRBTS-" in dn:
                return dn.split("/")[0]
    return ""

def get_lte_src_ips(xml_path):
    """
    Get srcIpAddress for LTE-Traffic (IPRT-8) and LTE-Signalling (IPRT-9).
    Find IPIF IPs that are NOT in the OAM gateway subnet.
    """
    tree = ET.parse(xml_path)
    root = tree.getroot()

    # Build maps: ipif_num → localIpAddr and ipif_num → vlanif_num
    ipif_ips  = {}
    ipif_vlan = {}
    for mo in root.iter():
        if _tag(mo) != "managedObject": continue
        dn  = mo.get("distName","")
        cls = mo.get("class","")
        m   = re.search(r'IPIF-(\d+)', dn)
        if not m: continue
        num = int(m.group(1))
        if "IPADDRESSV4" in cls:
            for child in mo:
                if _tag(child) == "p" and child.get("name") == "localIpAddr":
                    ipif_ips[num] = child.text or ""
        elif "IPIF" in cls and "IPADDRESSV4" not in cls:
            for child in mo:
                if _tag(child) == "p" and child.get("name") == "interfaceDN":
                    vm = re.search(r'VLANIF-(\d+)', child.text or "")
                    if vm: ipif_vlan[num] = int(vm.group(1))

    # Get OAM subnet to exclude non-LTE IPs
    oam_gw     = get_gateway_from_scf(xml_path)
    oam_subnet = ".".join(oam_gw.split(".")[:2]) if oam_gw else ""

    # Find non-OAM, non-loopback IPs sorted by VLAN number
    non_oam = []
    for vlan_num, ipif_num in sorted((v,k) for k,v in ipif_vlan.items()):
        ip = ipif_ips.get(ipif_num,"")
        if not ip or ip in ("0.0.0.0","10.10.10.1"): continue
        if oam_subnet and ip.startswith(oam_subnet + "."): continue
        non_oam.append((vlan_num, ip))

    lte_traffic_ip = non_oam[0][1] if len(non_oam) >= 1 else None
    lte_signal_ip  = non_oam[1][1] if len(non_oam) >= 2 else None

    return (lte_traffic_ip or "0.0.0.0"), (lte_signal_ip or "0.0.0.0")

def get_rtpol_prefix(xml_path):
    """Get srcIpPrefixLength from existing RTPOL items, preferring items 2-6."""
    tree = ET.parse(xml_path)
    root = tree.getroot()
    for mo in root.iter():
        if _tag(mo) != "managedObject": continue
        if "RTPOL" not in mo.get("distName",""): continue
        for lst in mo.iter():
            if _tag(lst) != "list": continue
            for item in lst:
                vals = {}
                for p in item:
                    if _tag(p) == "p":
                        vals[p.get("name","")] = p.text or ""
                # Use prefix from items 2-6 (NSA/SA) which are most reliable
                order = vals.get("orderNumber","")
                prefix = vals.get("srcIpPrefixLength","")
                if order in ("2","3","4","5","6") and prefix:
                    return prefix
    # Fallback: return first found
    for mo in root.iter():
        if _tag(mo) != "managedObject": continue
        if "RTPOL" not in mo.get("distName",""): continue
        for lst in mo.iter():
            if _tag(lst) != "list": continue
            for item in lst:
                for p in item:
                    if _tag(p) == "p" and p.get("name") == "srcIpPrefixLength":
                        return p.text or "27"
    return "27"

def compare(xml_routes, db_relations):
    xml_set = {(r["dest"], r["prefix"]): r for r in xml_routes}
    db_set  = {(r["dest"], r["prefix"]): r for r in db_relations}
    present, missing, extra = [], [], []
    for key, dr in db_set.items():
        if key in xml_set: present.append({**dr, "_xml_gw": xml_set[key]["gateway"]})
        else:              missing.append(dr)
    for key, xr in xml_set.items():
        if key not in db_set: extra.append(xr)
    return present, missing, extra


# ══════════════════════════════════════════
#  XML WRITERS
# ══════════════════════════════════════════
def write_missing_to_scf(xml_path, missing_routes):
    with open(xml_path, "r", encoding="utf-8") as f:
        content = f.read()
    iprt1_match = re.search(r'distName="[^"]*IPRT-1"', content)
    if not iprt1_match: return False, "IPRT-1 not found."
    static_open = content.find('name="staticRoutes"', iprt1_match.start())
    if static_open == -1: return False, "staticRoutes not found in IPRT-1."
    static_close = content.find("</list>", static_open)
    if static_close == -1: return False, "staticRoutes closing tag not found."
    def make_item(r):
        return (f"        <item>\n"
                f'          <p name="destinationIpPrefixLength">{r.get("prefix","0")}</p>\n'
                f'          <p name="destIpAddr">{r.get("dest","0.0.0.0")}</p>\n'
                f'          <p name="gateway">{r.get("gateway","0.0.0.0")}</p>\n'
                f'          <p name="preference">{r.get("pref","1")}</p>\n'
                f'          <p name="preSrcIpv4Addr">0.0.0.0</p>\n'
                f"        </item>")
    new_items = "\n" + "\n".join(make_item(r) for r in missing_routes)
    new_content = content[:static_close] + new_items + content[static_close:]
    with open(xml_path, "w", encoding="utf-8") as f:
        f.write(new_content)
    return True, f"{len(missing_routes)} relations written to IPRT-1."

def delete_iprt7_from_scf(xml_path):
    with open(xml_path, "r", encoding="utf-8") as f:
        content = f.read()
    pattern = r'<managedObject[^>]*distName="[^"]*IPRT-7"[^>]*>.*?</managedObject>'
    match = re.search(pattern, content, re.DOTALL)
    if not match: return False, "IPRT-7 not found in SCF."
    new_content = content[:match.start()] + content[match.end():]
    new_content = re.sub(r'\n\s*\n\s*\n', '\n\n', new_content)
    with open(xml_path, "w", encoding="utf-8") as f:
        f.write(new_content)
    return True, "IPRT-7 deleted."

def add_rtpol_to_scf(xml_path):
    with open(xml_path, "r", encoding="utf-8") as f:
        content = f.read()
    if 'orderNumber">8' in content or 'orderNumber">9' in content:
        return False, "RTPOL 8 or 9 already exists."
    dist_base  = get_distbase(xml_path)
    src_prefix = get_rtpol_prefix(xml_path)
    lte_traffic_ip, lte_signal_ip = get_lte_src_ips(xml_path)
    def rtpol_item(order, iprt_num, src_ip, prefix, label):
        return (f"        <item>\n"
                f'          <p name="orderNumber">{order}</p>\n'
                f'          <p name="routingTableDN">{dist_base}/TNLSVC-1/TNL-1/IPNO-1/IPRT-{iprt_num}</p>\n'
                f'          <p name="srcIpAddress">{src_ip}</p>\n'
                f'          <p name="srcIpPrefixLength">{prefix}</p>\n'
                f'          <p name="userLabel">{label}</p>\n'
                f"        </item>")
    new_items = ("\n" +
                 rtpol_item(8, 8, lte_traffic_ip, src_prefix, "LTE-Traffic") + "\n" +
                 rtpol_item(9, 9, lte_signal_ip,  src_prefix, "LTE-Signalling") + "\n      ")
    rtpol_match = re.search(r'name="routingPolicies"', content)
    if not rtpol_match: return False, "routingPolicies not found."
    close_pos = content.find("</list>", rtpol_match.start())
    if close_pos == -1: return False, "routingPolicies closing tag not found."
    new_content = content[:close_pos] + new_items + content[close_pos:]
    with open(xml_path, "w", encoding="utf-8") as f:
        f.write(new_content)
    return True, (f"RTPOL 8 & 9 added.\n"
                  f"  LTE-Traffic  → IPRT-8 | srcIP: {lte_traffic_ip}/{src_prefix}\n"
                  f"  LTE-Signalling → IPRT-9 | srcIP: {lte_signal_ip}/{src_prefix}")


# ══════════════════════════════════════════
#  MAIN APP
# ══════════════════════════════════════════
class SCFChecker(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("SCF Checker  v6.0  —  Nokia BTS")
        self.geometry("1280x820")
        self.minsize(1000, 680)
        self.configure(bg=BG_WIN)
        self.resizable(True, True)

        self.region_var   = tk.StringVar()
        self.solution_var = tk.StringVar()
        self.db_path_var  = tk.StringVar()
        self.scf_path_var = tk.StringVar()

        self.db_relations = []
        self.scf_routes   = []
        self.present      = []
        self.missing      = []
        self.extra        = []
        self.req_checks   = {}
        self.write_results = {}
        self.save_req_btn = None
        self.save_req_lbl = None

        self._apply_styles()
        self._build_ui()

    def _apply_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TCombobox",
                        fieldbackground=BG_INPUT, background=BG_INPUT,
                        foreground=TEXT1, selectbackground=ACCENT_LT,
                        selectforeground=TEXT1, bordercolor=BORDER, padding=4)
        style.map("TCombobox", fieldbackground=[("readonly", BG_INPUT)])
        style.configure("Main.Treeview",
                        background=BG_PANEL, foreground=TEXT1, rowheight=30,
                        fieldbackground=BG_PANEL, font=("Segoe UI",10), borderwidth=0)
        style.configure("Main.Treeview.Heading",
                        background="#e8edf5", foreground=TEXT1,
                        font=("Segoe UI",10,"bold"), relief="flat", padding=6)
        style.map("Main.Treeview",
                  background=[("selected", ACCENT_LT)],
                  foreground=[("selected", TEXT1)])
        style.configure("Sub.Treeview",
                        background=BG_PANEL, foreground=TEXT1, rowheight=26,
                        fieldbackground=BG_PANEL, font=("Segoe UI",9), borderwidth=0)
        style.configure("Sub.Treeview.Heading",
                        background="#e8edf5", foreground=TEXT2,
                        font=("Segoe UI",9,"bold"), relief="flat")
        style.map("Sub.Treeview",
                  background=[("selected", ACCENT_LT)],
                  foreground=[("selected", TEXT1)])

    # ══════════════════════════════════════
    def _build_ui(self):
        # Header
        hdr = tk.Frame(self, bg=ACCENT, pady=8, padx=16)
        hdr.pack(fill="x")
        tk.Label(hdr, text="⚡ SCF Checker",
                 font=("Segoe UI",14,"bold"), bg=ACCENT, fg=WHITE).pack(side="left")
        tk.Label(hdr, text="  v6.0  |  Nokia BTS Static Routes Validator",
                 font=F_SMALL, bg=ACCENT, fg="#cce8ff").pack(side="left")

        # ── Controls bar (vertical buttons) ──
        ctrl = tk.Frame(self, bg=BG_PANEL, pady=8, padx=10)
        ctrl.pack(fill="x")

        def vbtn(parent, num, title, widget_fn, num_color=ACCENT):
            """Create a vertical button block: number top, widget middle, label bottom"""
            f = tk.Frame(parent, bg=BG_PANEL, padx=6)
            f.pack(side="left")
            tk.Label(f, text=f" {num} ", font=F_NUM,
                     bg=num_color, fg=WHITE, padx=3).pack(anchor="center")
            widget_fn(f)
            tk.Label(f, text=title, font=("Segoe UI",8),
                     bg=BG_PANEL, fg=TEXT2).pack(anchor="center")
            return f

        def vsep():
            tk.Frame(ctrl, bg=BORDER, width=1, height=64).pack(side="left", padx=6)

        # 1 Region
        def mk_region(p):
            self.region_cb = ttk.Combobox(p, textvariable=self.region_var,
                                           values=REGIONS, state="readonly", width=8, font=F_BODY)
            self.region_cb.set("Select")
            self.region_cb.pack(pady=2)
            self.region_cb.bind("<<ComboboxSelected>>", lambda e: self._on_region_change())
        vbtn(ctrl, "1", "Region", mk_region)

        vsep()

        # 2 Solution
        def mk_sol(p):
            self.sol_cb = ttk.Combobox(p, textvariable=self.solution_var,
                                        values=SOLUTIONS, state="readonly", width=16, font=F_BODY)
            self.sol_cb.set("Select")
            self.sol_cb.pack(pady=2)
            self.sol_cb.bind("<<ComboboxSelected>>", lambda e: self._on_solution_change())
        vbtn(ctrl, "2", "Solution", mk_sol)

        vsep()

        # 3 DB file
        def mk_db(p):
            tk.Button(p, text="Browse…", font=F_SMALL, bg=ACCENT, fg=WHITE,
                      relief="flat", cursor="hand2", padx=10, pady=3,
                      command=self._browse_db).pack(pady=2)
            self.db_name_lbl = tk.Label(p, text="No file", font=("Segoe UI",8),
                                         bg=BG_PANEL, fg=TEXT3, width=16)
            self.db_name_lbl.pack()
        vbtn(ctrl, "3", "DB File (.xlsx)", mk_db, "#2e7d32")

        vsep()

        # 4 SCF file
        def mk_scf(p):
            tk.Button(p, text="Browse…", font=F_SMALL, bg=ACCENT, fg=WHITE,
                      relief="flat", cursor="hand2", padx=10, pady=3,
                      command=self._browse_scf).pack(pady=2)
            self.scf_name_lbl = tk.Label(p, text="No file", font=("Segoe UI",8),
                                          bg=BG_PANEL, fg=TEXT3, width=20)
            self.scf_name_lbl.pack()
        vbtn(ctrl, "4", "SCF File (.xml)", mk_scf, "#2e7d32")

        vsep()

        # 5 Clear SCF
        def mk_clrscf(p):
            tk.Button(p, text="✕ Clear SCF", font=F_SMALL,
                      bg=RED_BG, fg=RED, relief="flat", cursor="hand2",
                      padx=10, pady=3, command=self._clear_scf).pack(pady=2)
        vbtn(ctrl, "5", "Clear SCF", mk_clrscf, RED)

        vsep()

        # 6 RUN
        def mk_run(p):
            self.run_btn = tk.Button(p, text="▶  RUN",
                                      font=("Segoe UI",11,"bold"),
                                      bg=ACCENT, fg=WHITE, relief="flat",
                                      cursor="hand2", padx=18, pady=4,
                                      state="disabled", command=self._run)
            self.run_btn.pack(pady=2)
        vbtn(ctrl, "6", "Run Check", mk_run, "#0055aa")

        vsep()

        # 7 Write
        def mk_write(p):
            self.write_btn = tk.Button(p, text="✍  Write",
                                        font=("Segoe UI",10,"bold"),
                                        bg=GREEN_BG, fg=GREEN, relief="flat",
                                        cursor="hand2", padx=14, pady=4,
                                        state="disabled", command=self._write_missing)
            self.write_btn.pack(pady=2)
        vbtn(ctrl, "7", "Write Tasks", mk_write, "#1a6e38")

        vsep()

        # 8 Clear All
        def mk_clrall(p):
            tk.Button(p, text="Clear All", font=F_SMALL,
                      bg=BG_WIN, fg=TEXT2, relief="flat", cursor="hand2",
                      padx=10, pady=3, command=self._clear).pack(pady=2)
        vbtn(ctrl, "8", "Reset All", mk_clrall, TEXT3)

        # ── Requirements panel ──
        self.req_frame = tk.Frame(self, bg=BG_WIN)
        self.req_frame.pack(fill="x", padx=8)

        # ── Info bar ──
        self.info_bar = tk.Frame(self, bg=YELLOW_BG, padx=12)
        self.info_bar.pack(fill="x")
        self.info_lbl = tk.Label(self.info_bar, text="",
                                  font=F_SMALL, bg=YELLOW_BG, fg=YELLOW, anchor="w")
        self.info_lbl.pack(fill="x", pady=3)

        # ── Summary chips ──
        self.summary_frame = tk.Frame(self, bg=BG_WIN, padx=8, pady=4)
        self.summary_frame.pack(fill="x")

        # ── Main results table ──
        self._build_main_table()

        # ── Bottom panels (dynamic) — fixed height 220px ──
        self.bottom_frame = tk.Frame(self, bg=BG_WIN, height=220)
        self.bottom_frame.pack(fill="x", padx=8, pady=(4,8))
        self.bottom_frame.pack_propagate(False)  # keep fixed height

        # ── Status bar ──
        self.status_var = tk.StringVar(value="Ready — complete steps 1 to 4 then press RUN.")
        tk.Label(self, textvariable=self.status_var,
                 font=F_SMALL, bg=BG_CARD, fg=TEXT2,
                 anchor="w", padx=10).pack(fill="x", side="bottom")

        # ── Global copy: double-click any label copies its text ──
        def _global_copy(event):
            w = event.widget
            text = ""
            try:
                if isinstance(w, tk.Label):
                    text = w.cget("text")
                elif isinstance(w, tk.Entry):
                    try:
                        text = w.selection_get()
                    except:
                        text = w.get()
                elif isinstance(w, ttk.Combobox):
                    text = w.get()
                if text.strip():
                    self.clipboard_clear()
                    self.clipboard_append(text.strip())
            except: pass
        self.bind_all("<Double-Button-1>", _global_copy)

    # ─────────────────────────────────────
    def _build_main_table(self):
        outer = tk.Frame(self, bg=BG_WIN, padx=8)
        outer.pack(fill="both", expand=True)

        tf = tk.Frame(outer, bg=BG_WIN)
        tf.pack(fill="both", expand=True)

        vsb = ttk.Scrollbar(tf, orient="vertical")
        hsb = ttk.Scrollbar(tf, orient="horizontal")

        cols   = ("status","dest","prefix","gw_xml","gw_db","pref")
        hdrs   = ("  Status","  Destination IP","  Prefix",
                  "  Gateway (XML)","  Gateway (DB)","  Pref")
        widths = [170, 210, 90, 210, 210, 70]

        self.tree = ttk.Treeview(tf, columns=cols, show="headings",
                                  style="Main.Treeview",
                                  yscrollcommand=vsb.set,
                                  xscrollcommand=hsb.set)

        for col, hdr, w in zip(cols, hdrs, widths):
            self.tree.column(col, width=w, anchor="w", minwidth=60, stretch=True)
            self.tree.heading(col, text=hdr, anchor="w")

        self.tree.tag_configure("present", background="#f0faf4", foreground=GREEN)
        self.tree.tag_configure("missing", background=YELLOW_BG, foreground=YELLOW,
                                 font=("Segoe UI",10,"bold"))
        self.tree.tag_configure("extra",   background=ORANGE_BG, foreground=ORANGE)

        # Right-click copy
        def copy_row(event):
            sel = self.tree.selection()
            if not sel: return
            vals = self.tree.item(sel[0])["values"]
            text = "\t".join(str(v) for v in vals)
            self.clipboard_clear()
            self.clipboard_append(text)
        self.tree.bind("<Button-3>", copy_row)
        self.tree.bind("<Control-c>", copy_row)

        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tf.rowconfigure(0, weight=1)
        tf.columnconfigure(0, weight=1)

    # ─────────────────────────────────────
    def _build_requirements_panel(self):
        for w in self.req_frame.winfo_children():
            w.destroy()
        self.req_checks  = {}
        self.save_req_btn = None
        self.save_req_lbl = None

        region = self.region_var.get()
        sol    = self.solution_var.get()
        if region in ("","Select") or sol in ("","Select"):
            return

        outer = tk.Frame(self.req_frame, bg=BG_WIN)
        outer.pack(fill="x", pady=2)

        # ── Blue header bar ──
        panel = tk.Frame(outer, bg=BG_PANEL,
                         highlightbackground=BORDER, highlightthickness=1)
        panel.pack(side="left", fill="both", expand=True)

        hdr = tk.Frame(panel, bg=ACCENT, padx=10, pady=6)
        hdr.pack(fill="x")
        tk.Label(hdr, text="📋  Requirements", font=("Segoe UI",9,"bold"),
                 bg=ACCENT, fg=WHITE).pack(side="left")
        tk.Label(hdr, text=f"  {region}  /  {sol}",
                 font=F_SMALL, bg=ACCENT, fg="#cce8ff").pack(side="left")

        # Rows
        self.req_body = tk.Frame(panel, bg=BG_PANEL, padx=10, pady=6)
        self.req_body.pack(fill="x")
        self._render_req_rows()

        # ── Side buttons ──
        side = tk.Frame(outer, bg=BG_WIN, padx=4)
        side.pack(side="left", fill="y", padx=(4,0))

        tk.Button(side, text="＋  Add", font=("Segoe UI",8,"bold"),
                  bg=ACCENT, fg=WHITE, relief="flat", cursor="hand2",
                  padx=10, pady=4, command=self._req_add).pack(fill="x", pady=(0,3))
        tk.Button(side, text="✎  Modify", font=("Segoe UI",8,"bold"),
                  bg="#7a5200", fg=WHITE, relief="flat", cursor="hand2",
                  padx=10, pady=4, command=self._req_modify).pack(fill="x", pady=(0,3))
        tk.Button(side, text="✕  Delete", font=("Segoe UI",8,"bold"),
                  bg=RED, fg=WHITE, relief="flat", cursor="hand2",
                  padx=10, pady=4, command=self._req_delete).pack(fill="x", pady=(0,6))
        tk.Frame(side, bg=BORDER, height=1).pack(fill="x", pady=(0,6))
        self.save_req_btn = tk.Button(side, text="💾  Save", font=("Segoe UI",8,"bold"),
                  bg=GREEN, fg=WHITE, relief="flat", cursor="hand2",
                  padx=10, pady=4, command=self._req_save)
        self.save_req_btn.pack(fill="x")
        self.save_req_lbl = tk.Label(side, text="", font=("Segoe UI",8),
                                      bg=BG_WIN, fg=GREEN)
        self.save_req_lbl.pack()

    def _render_req_rows(self):
        for w in self.req_body.winfo_children():
            w.destroy()
        self.req_checks = {}
        reqs = get_requirements(self.region_var.get(), self.solution_var.get())
        for i, req in enumerate(reqs):
            row = tk.Frame(self.req_body, bg=BG_PANEL)
            row.pack(fill="x", pady=2)
            tk.Label(row, text=f" {i+1} ", font=("Segoe UI",7,"bold"),
                     bg=TEXT3, fg=WHITE, padx=2).pack(side="left", padx=(0,4))
            if req.get("auto"):
                lbl = tk.Label(row, text="○", font=("Segoe UI",11),
                               bg=BG_PANEL, fg=TEXT3, width=3)
                lbl.pack(side="left")
                tk.Label(row, text=req["label"], font=F_SMALL,
                         bg=BG_PANEL, fg=TEXT2).pack(side="left")
                self.req_checks[req["id"]] = {"type":"auto","lbl":lbl}
            else:
                var = tk.BooleanVar(value=req.get("checked", False))
                tk.Checkbutton(row, variable=var, bg=BG_PANEL,
                               activebackground=BG_PANEL, relief="flat",
                               cursor="hand2").pack(side="left")
                tk.Label(row, text=req["label"], font=F_SMALL,
                         bg=BG_PANEL, fg=TEXT1).pack(side="left")
                lbl = tk.Label(row, text="", font=("Segoe UI",11),
                               bg=BG_PANEL, fg=GREEN)
                lbl.pack(side="left", padx=6)
                self.req_checks[req["id"]] = {"type":"checkbox","var":var,"lbl":lbl}

    def _req_add(self):
        region, sol = self.region_var.get(), self.solution_var.get()
        if region in ("","Select") or sol in ("","Select"):
            messagebox.showwarning("Select First","Select Region and Solution first.")
            return
        dlg = tk.Toplevel(self); dlg.title("Add Requirement")
        dlg.geometry("460x200"); dlg.resizable(False,False)
        dlg.configure(bg=BG_WIN); dlg.grab_set()
        tk.Label(dlg, text="Add New Requirement", font=("Segoe UI",10,"bold"),
                 bg=BG_WIN, fg=TEXT1).pack(pady=(14,6))
        frm = tk.Frame(dlg, bg=BG_WIN, padx=20); frm.pack(fill="x")
        tk.Label(frm, text="Label:", font=F_SMALL, bg=BG_WIN, fg=TEXT2).pack(anchor="w")
        lbl_var = tk.StringVar()
        tk.Entry(frm, textvariable=lbl_var, font=F_BODY, width=46,
                 bg=BG_INPUT, relief="solid", bd=1).pack(fill="x", pady=(2,8))
        checked_var = tk.BooleanVar(value=True)
        tk.Checkbutton(frm, text="Checked by default", variable=checked_var,
                       bg=BG_WIN, font=F_SMALL, fg=TEXT1,
                       activebackground=BG_WIN).pack(anchor="w")
        def do_add():
            label = lbl_var.get().strip()
            if not label:
                messagebox.showwarning("Empty","Label cannot be empty.", parent=dlg); return
            key = (region, sol)
            new_id = f"custom_{len(SOLUTION_REQUIREMENTS.get(key,[]))}"
            if key not in SOLUTION_REQUIREMENTS: SOLUTION_REQUIREMENTS[key] = []
            SOLUTION_REQUIREMENTS[key].append({"id":new_id,"label":label,
                "auto":False,"trigger":"write","checked":checked_var.get()})
            dlg.destroy(); self._render_req_rows()
        tk.Button(dlg, text="  Add  ", font=("Segoe UI",9,"bold"),
                  bg=ACCENT, fg=WHITE, relief="flat", cursor="hand2",
                  padx=12, pady=4, command=do_add).pack(pady=10)

    def _req_modify(self):
        region, sol = self.region_var.get(), self.solution_var.get()
        key  = (region, sol)
        reqs = SOLUTION_REQUIREMENTS.get(key, [])
        if not reqs: messagebox.showinfo("None","No requirements to modify."); return
        dlg = tk.Toplevel(self); dlg.title("Modify Requirement")
        dlg.geometry("480x240"); dlg.resizable(False,False)
        dlg.configure(bg=BG_WIN); dlg.grab_set()
        tk.Label(dlg, text="Select requirement to modify:",
                 font=("Segoe UI",10,"bold"), bg=BG_WIN, fg=TEXT1).pack(pady=(14,6))
        frm = tk.Frame(dlg, bg=BG_WIN, padx=20); frm.pack(fill="x")
        labels = [f"{i+1}. {r['label']}" for i,r in enumerate(reqs)]
        sel_var = tk.StringVar(value=labels[0])
        cb = ttk.Combobox(frm, textvariable=sel_var, values=labels,
                          state="readonly", width=52, font=F_SMALL)
        cb.pack(fill="x", pady=(0,8))
        tk.Label(frm, text="New label:", font=F_SMALL, bg=BG_WIN, fg=TEXT2).pack(anchor="w")
        new_var = tk.StringVar()
        tk.Entry(frm, textvariable=new_var, font=F_BODY, width=52,
                 bg=BG_INPUT, relief="solid", bd=1).pack(fill="x", pady=(2,6))
        def on_sel(e=None): new_var.set(reqs[labels.index(sel_var.get())]["label"])
        cb.bind("<<ComboboxSelected>>", on_sel); on_sel()
        def do_mod():
            label = new_var.get().strip()
            if not label: messagebox.showwarning("Empty","Label cannot be empty.",parent=dlg); return
            SOLUTION_REQUIREMENTS[key][labels.index(sel_var.get())]["label"] = label
            dlg.destroy(); self._render_req_rows()
        tk.Button(dlg, text="  Save  ", font=("Segoe UI",9,"bold"),
                  bg="#7a5200", fg=WHITE, relief="flat", cursor="hand2",
                  padx=12, pady=4, command=do_mod).pack(pady=10)

    def _req_delete(self):
        region, sol = self.region_var.get(), self.solution_var.get()
        key  = (region, sol)
        reqs = SOLUTION_REQUIREMENTS.get(key, [])
        if not reqs: messagebox.showinfo("None","No requirements to delete."); return
        dlg = tk.Toplevel(self); dlg.title("Delete Requirement")
        dlg.geometry("480x200"); dlg.resizable(False,False)
        dlg.configure(bg=BG_WIN); dlg.grab_set()
        tk.Label(dlg, text="Select requirement to delete:",
                 font=("Segoe UI",10,"bold"), bg=BG_WIN, fg=TEXT1).pack(pady=(14,6))
        frm = tk.Frame(dlg, bg=BG_WIN, padx=20); frm.pack(fill="x")
        labels = [f"{i+1}. {r['label']}" for i,r in enumerate(reqs)]
        sel_var = tk.StringVar(value=labels[0])
        ttk.Combobox(frm, textvariable=sel_var, values=labels,
                     state="readonly", width=52, font=F_SMALL).pack(fill="x", pady=(0,8))
        def do_del():
            if not messagebox.askyesno("Confirm",
                    f"Delete:\n\"{reqs[labels.index(sel_var.get())]['label']}\"?",
                    parent=dlg): return
            SOLUTION_REQUIREMENTS[key].pop(labels.index(sel_var.get()))
            dlg.destroy(); self._render_req_rows()
        tk.Button(dlg, text="  Delete  ", font=("Segoe UI",9,"bold"),
                  bg=RED, fg=WHITE, relief="flat", cursor="hand2",
                  padx=12, pady=4, command=do_del).pack(pady=10)

    def _req_save(self):
        ok, msg = save_requirements()
        if ok:
            if self.save_req_lbl and self.save_req_btn:
                try:
                    self.save_req_lbl.config(text=" ✔ Saved", fg=GREEN, bg=GREEN_BG)
                    self.save_req_btn.config(bg="#155a2a")
                    def _reset():
                        try:
                            self.save_req_lbl.config(text="", bg=BG_WIN)
                            self.save_req_btn.config(bg=GREEN)
                        except: pass
                    self.after(2000, _reset)
                except: messagebox.showinfo("Saved","Requirements saved.")
            else:
                messagebox.showinfo("Saved","Requirements saved.")
        else:
            messagebox.showerror("Error", f"Could not save:\n{msg}")

    # ─────────────────────────────────────
    def _build_bottom_panels(self):
        for w in self.bottom_frame.winfo_children():
            w.destroy()

        region = self.region_var.get()
        sol    = self.solution_var.get()
        reqs   = get_requirements(region, sol)
        write_reqs = [r for r in reqs if not r.get("auto")]
        if not write_reqs: return

        n = len(write_reqs)
        for i, req in enumerate(write_reqs):
            task_id = req["id"]
            result  = self.write_results.get(task_id)
            ok      = result[0] if result else None
            msg     = result[1] if result else ""

            hdr_color = GREEN_BG if ok is True else (RED_BG if ok is False else BG_CARD)
            hdr_fg    = GREEN    if ok is True else (RED    if ok is False else TEXT2)
            icon = "✔" if ok is True else ("✘" if ok is False else "○")

            panel = tk.Frame(self.bottom_frame, bg=BG_PANEL,
                             highlightbackground=BORDER, highlightthickness=1)
            panel.grid(row=0, column=i, sticky="nsew",
                       padx=(0 if i==0 else 3, 0))
            self.bottom_frame.columnconfigure(i, weight=1)
            self.bottom_frame.rowconfigure(0, weight=1)

            # Header
            hdr = tk.Frame(panel, bg=hdr_color, padx=8, pady=4)
            hdr.pack(fill="x")
            tk.Label(hdr, text=f"{icon}  {req['label']}",
                     font=("Segoe UI",8,"bold"), bg=hdr_color, fg=hdr_fg,
                     wraplength=350, justify="left").pack(anchor="w")

            body = tk.Frame(panel, bg=BG_PANEL)
            body.pack(fill="both", expand=True)

            if task_id == "add_iprt1":
                self._panel_iprt1_v2(body)
            elif task_id == "delete_iprt7":
                self._panel_iprt7_v2(body, ok, msg)
            elif task_id == "add_rtpol":
                self._panel_rtpol_v2(body, ok, msg)
            else:
                self._panel_generic(body, ok, msg)

    def _selectable_text(self, parent, text, fg=None, font=None, bg=None):
        """Create a selectable (copyable) text entry-style label"""
        bg    = bg    or BG_PANEL
        fg    = fg    or TEXT1
        font  = font  or ("Segoe UI", 9)
        e = tk.Entry(parent, font=font, fg=fg, bg=bg,
                     relief="flat", bd=0,
                     readonlybackground=bg,
                     state="readonly")
        e.insert(0, text)
        e.config(state="readonly")
        e.pack(fill="x", padx=6, pady=1)
        return e

    def _panel_iprt1_v2(self, parent):
        # Summary chips
        chips = tk.Frame(parent, bg=BG_PANEL, padx=4, pady=4)
        chips.pack(fill="x")
        tk.Label(chips, text=f"✔ {len(self.present)} present",
                 font=("Segoe UI",9,"bold"), bg=GREEN_BG, fg=GREEN,
                 padx=6, pady=2).pack(side="left", padx=2)
        tk.Label(chips, text=f"⚠ {len(self.missing)} missing",
                 font=("Segoe UI",9,"bold"), bg=YELLOW_BG, fg=YELLOW,
                 padx=6, pady=2).pack(side="left", padx=2)

        if not self.missing:
            tk.Label(parent, text="All relations present ✔",
                     font=("Segoe UI",9), bg=GREEN_BG, fg=GREEN,
                     padx=8, pady=4).pack(fill="x", padx=4, pady=2)
            return

        # Scrollable list of missing routes
        lbl = tk.Label(parent, text="Missing routes written to IPRT-1:",
                       font=("Segoe UI",8,"bold"), bg=BG_PANEL, fg=TEXT2)
        lbl.pack(anchor="w", padx=6, pady=(4,0))

        outer = tk.Frame(parent, bg=BG_PANEL)
        outer.pack(fill="both", expand=True, padx=4, pady=2)
        vsb = ttk.Scrollbar(outer, orient="vertical")
        vsb.pack(side="right", fill="y")
        canvas = tk.Canvas(outer, bg=BG_PANEL, highlightthickness=0,
                           yscrollcommand=vsb.set)
        canvas.pack(side="left", fill="both", expand=True)
        vsb.config(command=canvas.yview)
        inner = tk.Frame(canvas, bg=BG_PANEL)
        cw = canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(cw, width=e.width))

        for r in self.missing:
            row = tk.Frame(inner, bg=YELLOW_BG)
            row.pack(fill="x", pady=1)
            txt = f"  {r.get('dest','')}/{r.get('prefix','')}  →  {r.get('gateway','')}"
            e = tk.Entry(row, font=F_MONO, fg=YELLOW, bg=YELLOW_BG,
                         relief="flat", bd=0, readonlybackground=YELLOW_BG,
                         state="readonly")
            e.insert(0, txt)
            e.config(state="readonly")
            e.pack(fill="x", padx=4, pady=1)

    def _panel_iprt7_v2(self, parent, ok, msg):
        if ok is None:
            tk.Label(parent, text="Not executed", font=F_SMALL,
                     bg=BG_PANEL, fg=TEXT3, padx=8, pady=8).pack(anchor="w")
            return

        if not ok:
            tk.Label(parent, text=f"✘ {msg}", font=F_SMALL,
                     bg=RED_BG, fg=RED, padx=8, pady=6).pack(fill="x")
            return

        tk.Label(parent, text="✔  IPRT-7 removed — IPv4 Routing now:",
                 font=("Segoe UI",9,"bold"), bg=GREEN_BG, fg=GREEN,
                 padx=8, pady=4).pack(fill="x")

        # Show remaining IPRT list (like Nokia UI)
        iprts = [
            ("IPRT-1", "NSA-OAM"),
            ("IPRT-2", "NSA-Traffic"),
            ("IPRT-3", "NSA-Signalling"),
            ("IPRT-4", "SA-N3"),
            ("IPRT-5", "SA-N2"),
            ("IPRT-6", "SA-Xn"),
            ("IPRT-8", "LTE-Traffic"),
            ("IPRT-9", "LTE-Signalling"),
        ]
        list_frame = tk.Frame(parent, bg=BG_PANEL, padx=6, pady=4)
        list_frame.pack(fill="both", expand=True)

        for iprt_id, label in iprts:
            row = tk.Frame(list_frame, bg=BG_PANEL,
                           highlightbackground=BORDER, highlightthickness=1)
            row.pack(fill="x", pady=1)
            tk.Label(row, text="▶", font=("Segoe UI",9),
                     bg=BG_PANEL, fg=TEXT3, width=2).pack(side="left", padx=4)
            e = tk.Entry(row, font=("Segoe UI",9), fg=TEXT1, bg=BG_PANEL,
                         relief="flat", bd=0, readonlybackground=BG_PANEL,
                         state="readonly")
            e.insert(0, f"{iprt_id}  {label}")
            e.config(state="readonly")
            e.pack(side="left", fill="x", expand=True, padx=4, pady=3)

    def _panel_rtpol_v2(self, parent, ok, msg):
        if ok is None:
            tk.Label(parent, text="Not executed", font=F_SMALL,
                     bg=BG_PANEL, fg=TEXT3, padx=8, pady=8).pack(anchor="w")
            return
        if not ok:
            tk.Label(parent, text=f"✘ {msg}", font=F_SMALL,
                     bg=RED_BG, fg=RED, padx=8, pady=6).pack(fill="x")
            return

        tk.Label(parent, text="✔  Routing Policies added:",
                 font=("Segoe UI",9,"bold"), bg=GREEN_BG, fg=GREEN,
                 padx=8, pady=4).pack(fill="x")

        body = tk.Frame(parent, bg=BG_PANEL, padx=6, pady=4)
        body.pack(fill="both", expand=True)

        for line in msg.split("\n"):
            line = line.strip()
            if not line: continue
            bg = GREEN_BG if ("Traffic" in line or "Signal" in line or "✔" in line) else BG_PANEL
            fg = GREEN    if ("Traffic" in line or "Signal" in line or "✔" in line) else TEXT1
            e = tk.Entry(body, font=F_MONO, fg=fg, bg=bg,
                         relief="flat", bd=0, readonlybackground=bg,
                         state="readonly")
            e.insert(0, f"  {line}")
            e.config(state="readonly")
            e.pack(fill="x", pady=1)

    def _panel_generic(self, parent, ok, msg):
        color = GREEN_BG if ok else (RED_BG if ok is False else BG_PANEL)
        fg    = GREEN    if ok else (RED    if ok is False else TEXT2)
        for line in (msg or "Not executed").split("\n"):
            e = tk.Entry(parent, font=F_SMALL, fg=fg, bg=color,
                         relief="flat", bd=0, readonlybackground=color,
                         state="readonly")
            e.insert(0, f"  {line.strip()}")
            e.config(state="readonly")
            e.pack(fill="x", padx=6, pady=1)

    # ══════════════════════════════════════
    def _on_region_change(self):
        self._check_can_run()
        self._build_requirements_panel()

    def _on_solution_change(self):
        self._check_sheet()
        self._check_can_run()
        self._build_requirements_panel()

    def _browse_db(self):
        p = filedialog.askopenfilename(
            title="Select DB File (Excel)",
            filetypes=[("Excel files","*.xlsx *.xls"),("All","*.*")])
        if not p: return
        self.db_path_var.set(p)
        fname = os.path.basename(p)
        self.db_name_lbl.config(
            text=fname[:18]+"…" if len(fname)>18 else fname, fg=GREEN)
        self._check_sheet()
        self._check_can_run()

    def _browse_scf(self):
        p = filedialog.askopenfilename(
            title="Select SCF File (XML)",
            filetypes=[("XML files","*.xml"),("All","*.*")])
        if not p: return
        self.scf_path_var.set(p)
        fname = os.path.basename(p)
        self.scf_name_lbl.config(
            text=fname[:22]+"…" if len(fname)>22 else fname, fg=GREEN)
        self._check_can_run()

    def _check_sheet(self):
        db, sol = self.db_path_var.get(), self.solution_var.get()
        if not db or not sol or sol == "Select":
            self.info_lbl.config(text=""); return
        if not EXCEL_SUPPORT: return
        try:
            wb = openpyxl.load_workbook(db, read_only=True)
            sheets = wb.sheetnames; wb.close()
            matched = any(s.strip().upper() == sol.strip().upper() for s in sheets)
            if matched:
                self.info_bar.config(bg="#e8f5ec")
                self.info_lbl.config(text=f"✔  Sheet '{sol}' found", bg="#e8f5ec", fg=GREEN)
            else:
                self.info_bar.config(bg=YELLOW_BG)
                self.info_lbl.config(
                    text=f"⚠  Sheet '{sol}' not found  |  Available: {', '.join(sheets)}",
                    bg=YELLOW_BG, fg=YELLOW)
        except Exception as e:
            self.info_bar.config(bg=RED_BG)
            self.info_lbl.config(text=f"✕  {e}", bg=RED_BG, fg=RED)

    def _check_can_run(self):
        ok = (self.region_var.get() not in ("","Select") and
              self.solution_var.get() not in ("","Select") and
              bool(self.db_path_var.get()) and
              bool(self.scf_path_var.get()))
        self.run_btn.config(state="normal" if ok else "disabled",
                             bg=ACCENT if ok else "#c0cfe0",
                             fg=WHITE if ok else TEXT3)

    # ══════════════════════════════════════
    def _run(self):
        region, solution = self.region_var.get(), self.solution_var.get()
        db_path, scf_path = self.db_path_var.get(), self.scf_path_var.get()

        self.status_var.set("Reading DB…"); self.update_idletasks()
        self.db_relations = read_relations_from_excel(db_path, solution)
        if not self.db_relations:
            self.status_var.set("No relations loaded."); return

        self.status_var.set("Parsing SCF…"); self.update_idletasks()
        try:
            self.scf_routes = parse_scf_iprt1(scf_path)
        except Exception as e:
            messagebox.showerror("SCF Error", str(e)); return

        self.status_var.set("Comparing…"); self.update_idletasks()
        scf_gw = get_gateway_from_scf(scf_path)
        for r in self.db_relations:
            if r.get("gateway") in ("FROM_SCF","","0.0.0.0"):
                r["gateway"] = scf_gw

        self.present, self.missing, self.extra = compare(self.scf_routes, self.db_relations)
        self._render_results(region, solution)

        if "check_scf" in self.req_checks:
            self.req_checks["check_scf"]["lbl"].config(text=" ✔", fg=GREEN)

        self.write_btn.config(state="normal", bg=GREEN_BG, fg=GREEN)
        self.status_var.set(
            f"Done · {region}/{solution} · "
            f"✔{len(self.present)} present · ⚠{len(self.missing)} missing · ⊕{len(self.extra)} extra")

    def _render_results(self, region, solution):
        for item in self.tree.get_children(): self.tree.delete(item)
        for w in self.summary_frame.winfo_children(): w.destroy()

        for text, bg, fg in [
            (f"  ✔  {len(self.present)} Present  ", GREEN_BG, GREEN),
            (f"  ⚠  {len(self.missing)} Missing  ", YELLOW_BG, YELLOW),
            (f"  ⊕  {len(self.extra)} Extra  ",     ORANGE_BG, ORANGE),
        ]:
            tk.Label(self.summary_frame, text=text, font=("Segoe UI",9,"bold"),
                     bg=bg, fg=fg, padx=8, pady=3).pack(side="left", padx=3)
        tk.Label(self.summary_frame, text=f"  IPRT-1  ·  {region} / {solution}",
                 font=F_MONO, bg=BG_WIN, fg=ACCENT).pack(side="left", padx=10)

        for r in self.present:
            self.tree.insert("","end", values=(
                "✔  Done", r.get("dest",""), "/"+r.get("prefix",""),
                r.get("_xml_gw",""), r.get("gateway",""), r.get("pref","1")
            ), tags=("present",))
        for r in self.missing:
            self.tree.insert("","end", values=(
                "⚠  Missing", r.get("dest",""), "/"+r.get("prefix",""),
                "— not in SCF —", r.get("gateway",""), r.get("pref","1")
            ), tags=("missing",))
        for r in self.extra:
            self.tree.insert("","end", values=(
                "⊕  Extra", r.get("dest",""), "/"+r.get("prefix",""),
                r.get("gateway",""), "— not in DB —", r.get("pref","1")
            ), tags=("extra",))

    # ══════════════════════════════════════
    def _write_missing(self):
        scf_path = self.scf_path_var.get()
        if not scf_path:
            messagebox.showwarning("No SCF","Load an SCF file first."); return

        do_iprt1 = self.req_checks.get("add_iprt1",{}).get("var", tk.BooleanVar()).get()
        do_del7  = self.req_checks.get("delete_iprt7",{}).get("var", tk.BooleanVar()).get()
        do_rtpol = self.req_checks.get("add_rtpol",{}).get("var", tk.BooleanVar()).get()

        if not any([do_iprt1, do_del7, do_rtpol]):
            messagebox.showwarning("Nothing selected","Check at least one task in Requirements."); return

        self.write_results = {}
        self.status_var.set("Writing…"); self.update_idletasks()
        results_summary = []

        if do_iprt1:
            ok, msg = write_missing_to_scf(scf_path, self.missing) if self.missing else (True, "No missing relations — already complete.")
            self.write_results["add_iprt1"] = (ok, msg)
            results_summary.append(f"{'✔' if ok else '✘'} IPRT-1: {msg}")
            if "add_iprt1" in self.req_checks:
                self.req_checks["add_iprt1"]["lbl"].config(text=" ✔" if ok else " ✘", fg=GREEN if ok else RED)

        if do_del7:
            ok, msg = delete_iprt7_from_scf(scf_path)
            self.write_results["delete_iprt7"] = (ok, msg)
            results_summary.append(f"{'✔' if ok else '✘'} IPRT-7: {msg}")
            if "delete_iprt7" in self.req_checks:
                self.req_checks["delete_iprt7"]["lbl"].config(text=" ✔" if ok else " ✘", fg=GREEN if ok else RED)

        if do_rtpol:
            ok, msg = add_rtpol_to_scf(scf_path)
            self.write_results["add_rtpol"] = (ok, msg)
            results_summary.append(f"{'✔' if ok else '✘'} RTPOL: {msg.split(chr(10))[0]}")
            if "add_rtpol" in self.req_checks:
                self.req_checks["add_rtpol"]["lbl"].config(text=" ✔" if ok else " ✘", fg=GREEN if ok else RED)

        messagebox.showinfo("Write Complete", "\n".join(results_summary))
        self.status_var.set("Write complete. Re-running check…")
        self.update_idletasks()
        self._run()
        self._build_bottom_panels()
        self.update_idletasks()

    # ══════════════════════════════════════
    def _clear_scf(self):
        self.scf_path_var.set("")
        self.scf_name_lbl.config(text="No file", fg=TEXT3)
        self.scf_routes = []
        self.present = self.missing = self.extra = []
        self.write_results = {}
        for item in self.tree.get_children(): self.tree.delete(item)
        for w in self.summary_frame.winfo_children(): w.destroy()
        for w in self.bottom_frame.winfo_children(): w.destroy()
        self.run_btn.config(state="disabled", bg="#c0cfe0", fg=TEXT3)
        self.write_btn.config(state="disabled", bg=BG_CARD, fg=TEXT3)
        self.status_var.set("SCF cleared.")
        self._check_can_run()

    def _clear(self):
        self.region_var.set("Select")
        self.solution_var.set("Select")
        self.db_path_var.set("")
        self.scf_path_var.set("")
        self.db_name_lbl.config(text="No file", fg=TEXT3)
        self.scf_name_lbl.config(text="No file", fg=TEXT3)
        self.info_lbl.config(text="")
        self.info_bar.config(bg=YELLOW_BG)
        self.db_relations = []
        self.scf_routes   = []
        self.present = self.missing = self.extra = []
        self.write_results = {}
        self.req_checks = {}
        for item in self.tree.get_children(): self.tree.delete(item)
        for w in self.summary_frame.winfo_children(): w.destroy()
        for w in self.req_frame.winfo_children(): w.destroy()
        for w in self.bottom_frame.winfo_children(): w.destroy()
        self.run_btn.config(state="disabled", bg="#c0cfe0", fg=TEXT3)
        self.write_btn.config(state="disabled", bg=BG_CARD, fg=TEXT3)
        self.status_var.set("Cleared. Ready.")


# ══════════════════════════════════════════
if __name__ == "__main__":
    app = SCFChecker()
    app.mainloop()
